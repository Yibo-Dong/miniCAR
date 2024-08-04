import json5
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Color

def getbasename(path):
    filename = os.path.basename(path)
    basename = filename.split(".")[0]
    return basename

vbs = set()
breakthrough = dict()
res = ""

hard = {'6s158', 'intel032', 'oski15a01b44s', 'beemldfilt7b1', '6s100', 'oski15a01b72s', 'oski15a14b20s', 'oski15a01b66s', 'oski15a10b08s', '6s128', 'oski15a14b04s', 'intel027', '6s340rb63', '6s161', 'oski15a10b06s', 'oski15a01b54s', '6s45', 'beembkry8b1', 'beemfwt4b1', 'oski15a01b34s', 'oski15a14b28s', '6s29', 'intel014', 'beemfwt5b1', '6s33', 'oski15a01b42s', 'oc8051gm49acc', '6s376r', 'oski1rub00i', '6s149', 'oski15a01b76s', 'oski15a01b14s', '6s316b421', '6s42', 'bobpcihm', '6s398b09', 'oski1rub10i', 'oski15a01b58s', 'oski15a01b08s', 'oc8051gm3bacc', '6s160', 'oski1rub02i', 'oski15a10b04s', 'oski15a08b10s', '6s171', 'oc8051gmd7acc', '6s267rb3', '6s185', 'bob12s02', 'beemskbn2b1', '6s280r', 'beemhanoi4b1', '6s329rb19', 'oc8051gm88iram', '6s398b16', '6s268r', '6s382r', 'intel016', 'oski15a14b06s', 'oski15a10b16s', 'oski15a01b48s', '6s195', '6s190', 'oski15a01b32s', 'beemcmbrdg7f2', 'oski15a01b70s', 'nusmvdme216', 'beemandrsn6b1', '6s22', 'oc8051gma4pc', 'oski15a01b02s', '6s342rb122', 'beempgmprot1f2', '6s274r', 'beemlifts3b1', '6s37', 'oski15a01b64s', 'intel013', 'oski15a01b50s', '6s179', '6s399b02', '6s163', 'oski1rub08i', 'oski15a01b78s', 'oski15a01b28s', 'intel067', '6s517rb0', '6s191', 'beemfwt5b3', '6s266rb2', 'oski15a01b52s', '6s39', 'oski15a01b60s', '6s148', 'oski1rub09i', 'oc8051gm43acc', 'oc8051gm06iram', 'oski15a01b68s', 'oski15a01b74s', '6s316b460', 'oski15a01b38s', '6s188', 'oski15a08b12s', 'oski15a01b00s', '6s279r', 'oski15a14b30s', '6s392r', 'intel065', 'oski15a07b0s', '6s105', 'oc8051gm63iram', 'beemloyd3b1', 'oski15a01b36s', '6s23', 'oski1rub01i', 'oski15a14b02s', 'oc8051gmbfpc', 'oski15a01b30s', '6s365r', '6s516r', '6s340rb27', '6s514r', 'oski15a01b12s', 'oski15a07b2s', 'oski15a01b62s', 'oski15a10b02s', '6s44', 'oski15a08b02s', '6s36', '6s377r', 'oski15a01b40s', 'oski15a14b26s', 'oski15a01b26s', 'oski15a10b10s', 'intel012', 'beemkrebs4b1', 'intel028', '6s399b03', '6s329rb20', 'oski15a01b18s', '6s186', 'oski15a07b4s', '6s24', '6s341r', 'oski15a08b04s'}

not_safe = {'6s160', 'intel012', 'oski15a08b09s', 'oski15a08b12s', 'oski15a10b09s', '6s514r', 'oski15a01b14s', '6s37', '6s401rb051', '6s341r', '6s329rb20', 'oc8051gm06iram', 'beemfwt5b1', '6s392r', '6s374b029', 'beembkry8b1', '6s20', 'oski15a14b29s', 'oski15a14b13s', 'intel040', 'oski15a08b16s', '6s339rb19', 'irstdme6', 'oski15a01b25s', 'oski15a01b48s', 'oski15a01b07s', 'bc57sensorsp0', 'intel035', 'oc8051gm63iram', 'oc8051gm88iram', '6s342rb122', '6s187', 'oski15a08b06s', 'oski15a01b01s', 'oski15a01b52s', 'bob9234spec5neg', 'oski15a01b43s', '6s398b09', '6s177', 'oski3b3i', 'oski15a10b10s', '6s274r', 'oski15a01b68s', 'oski15a01b36s', 'bobsynth13', '6s42', 'oski15a01b59s', '6s54', 'beemprdcell2f1', 'oski15a01b18s', 'oski15a07b1s', 'oski15a01b02s', '6s399b02', '6s319r', 'intel016', 'oski15a14b33s', 'oski15a01b42s', '6s279r', '6s24', 'beempgmprot1f2', 'dme6p1neg', '6s340rb63', '6s380b511', 'intel039', '6s301rb106', 'intel028', 'intel043', 'bj08amba2g4f3', '6s31', '6s119', 'oski15a10b03s', 'oski15a01b39s', 'oski15a01b50s', '6s148', 'intel038', '6s284rb1', 'oski15a14b28s', 'oski2b1i', 'prodcellp0neg', 'beemadd4b1', 'intel042', '139443p5', 'oski15a10b13s', '6s280r', 'prodcellp1', 'bobtuttt', '6s340rb27', 'intel044', 'oski15a01b58s', 'oski3b0i', '6s33', 'oski15a08b17s', 'intel032', '6s402rb0342', 'bobsynthor', 'brpp1', 'oski15a01b35s', 'oski15a08b15s', '6s365r', 'bobmiterbm1or', 'intel013', 'irstdme5', '6s216rb0', 'oski15a10b11s', '6s316b460', 'oski15a01b00s', '6s195', '6s171', 'intel046', 'oski15a01b75s', '6s186', 'oski2b4i', 'intel045', 'oski15a01b60s', 'oski15a14b26s', 'oc8051gma4pc', 'mutexp0', 'oc8051gm49acc', 'oski15a01b72s', 'oski15a01b76s', '6s12', 'csmacdp2', 'oski15a14b23s', 'oski15a07b0s', 'oski15a14b03s', 'oski15a08b08s', 'beemkrebs4b1', 'oski1rub01i', 'oski15a01b57s', 'oski1rub00i', '6s215rb0', 'oski15a10b05s', '6s339rb22', '6s161', 'oski15a01b73s', 'oski15a08b01s', '6s134', '6s158', '6s309b034', 'beempgsol5b1', '6s41', 'bobsynthetic2', 'oski15a01b65s', 'oski15a01b79s', 'oski2ub0i', 'oski15a01b53s', 'nusmvdme216', 'oski15a01b33s', 'oc8051gmbfpc', '6s23', 'oski15a01b74s', 'oski15a08b04s', 'bob12s02', 'oski15a01b77s', 'oski1rub02i', 'oski15a08b07s', 'oski15a08b05s', 'oski15a01b71s', 'oski15a14b06s', 'productioncellp0', '6s10', 'bobpcihm', '6s207rb28', 'mentorbm1and', '6s351rb02', '6s399b03', 'oski15a01b51s', '6s128', 'oski15a01b19s', 'oski15a01b27s', 'oski15a01b55s', 'intel047', '6s179', 'oski15a01b70s', 'prodcellp2neg', 'oski15a10b07s', 'oski15a01b08s', '6s398b16', 'oski15a10b04s', 'oc8051gmd7acc', '6s210b037', 'oski15a14b01s', '6s268r', 'oski15a01b44s', 'oski15a08b03s', 'oc8051gm43acc', '6s100', 'oski15a01b31s', 'oski15a01b13s', 'oski15a08b13s', 'oski15a14b04s', '6s266rb2', '6s29', '6s190', 'oski15a01b78s', 'intel065', '6s329rb19', 'oski15a01b03s', 'oski15a01b63s', '6s218b2950', 'intel017', 'oski15a01b17s', '6s517rb0', 'oski15a14b31s', '6s149', 'viselevatorp2', 'beemfwt5b3', '6s188', '6s382r', 'oski15a01b69s', '6s184', 'oski15a01b29s', '6s39', 'oski15a01b09s', '6s191', 'texasifetch1p8', '6s122', '6s13', 'oski15a01b38s', '139442p1', 'oski1rub10i', 'oski1rub09i', 'oski15a07b3s', '6s357r', 'bc57sensorsp1', '6s376r', 'srg5ptimo', 'oski15a10b01s', 'oski15a14b27s', 'oski15a14b07s', 'oski15a14b05s', '6s105', 'oski1rub08i', 'beemfrogs4f1', 'oski15a08b14s', '6s44', 'pdtswvibs8x8p0', 'oski15a14b21s', '6s309b046', '6s367r', 'oski15a07b4s', 'dme6p1', 'beemloyd3b1', 'oski15a01b45s', 'beemcmbrdg7f2', 'oski15a01b15s', '6s45', '6s389b02', 'intel067', 'ringp0', 'abp4p2ff', 'oski15a01b66s', 'oski15a01b47s', 'oski15a08b02s', 'oski15a01b37s', '6s316b421', 'oski15a14b25s', 'oski15a14b09s', 'intel027', '6s350rb35', 'oski15a01b26s', '6s320rb1', 'oski15a14b19s', 'bc57sensorsp2', 'oski15a10b17s', 'beemhanoi4b1', 'oski15a10b16s', 'oski15a14b02s', '6s299b685', 'oski15a14b17s', '6s22', 'intel048', 'oski15a01b41s', '6s16', 'counterp0', 'oski15a01b30s', 'oski15a08b11s', 'oski15a01b34s', 'oski15a10b02s', '6s351rb15', 'intel036', 'beembrptwo6b1', 'beemandrsn6b1', 'oski15a01b21s', 'oski15a01b40s', '6s377r', 'intel014', 'oski15a10b15s', 'oski2b5i', 'bc57sensorsp1neg', 'oski15a01b61s', 'oski15a01b49s', 'oski15a10b08s', 'beemldfilt7b1', 'oski15a14b15s', 'oski15a01b67s', 'pcip1', 'oski15a01b62s', 'beemfwt4b1', 'oski15a01b54s', '6s267rb3', 'oc8051topo08', 'oski15a07b5s', 'texasparsesysp3', '139444p22', 'oc8051gm3bacc', '6s7', 'beemextnc3f1', 'oski15a14b30s', 'bob12s03', 'oski15a10b14s', 'beemlifts3b1', '6s36', 'intel009', 'beemlmprt8f1', 'bob9234spec4neg', 'bj08vendingcycle', 'beemskbn2b1', '6s185', 'oski15a10b06s', '6s163', 'prodcellp4', 'oski15a01b12s', 'oski15a07b2s', 'bob9234spec6neg', 'oski15a01b23s', 'oski15a01b64s', '6s350rb46', 'oski2b3i', 'oski15a01b28s', 'bobtuint24', 'oski15a01b32s', 'oski15a08b10s', '6s516r', 'oski15a14b11s', '6s320rb0', 'oski15a14b20s'}

class info:
    global vbs
    def __init__(self, path):
        try:
            with open(path, "r") as fp:
                self.name = getbasename(path)
                try:
                    content = json5.load(fp)
                except Exception as e:
                    print(f"error in loading log {path} : {e}")
                    content = fp.read()
                if len(content) == 0:
                    self.status = "missed"
                else:
                    self.color = "black"
                    self.status = content["Status"]
                    if self.status == "cex found":
                        vbs.add(self.name)
                        if self.name in hard:
                            strategy = path.split("/")[-2]
                            if breakthrough.get(strategy) is None:
                                breakthrough[strategy] = [self.name]
                            else:
                                breakthrough[strategy].append(self.name)
                            self.color = "red"
                    original = content["Original main solver SAT Calls"]
                    self.time_original = original["Total Time"]
                    self.count_original = original["Total Count"]
                    self.count_original_success = original["Success"]
                    self.count_original_failed = original["Failed"]
                    self.count_tried_before = original["Tried before"]
                    self.count_original_uc_len = original["UC length avergage"]
                    conv = content["Convergence main solver SAT Calls"]
                    self.time_conv = conv["Total Time"]
                    self.count_conv = conv["Total Count"]
                    self.count_conv_shorter = conv["shorter UC"]
                    self.count_conv_uc_len = conv["UC length avergage"]
                    if content.get("Implication") != None:
                        impl = content["Implication"]
                        self.time_imply = impl["Total Time"]
                        self.count_imply = impl["Group Count"]
                        self.impDecision = impl["Decision"]
                    self.count_rounds = content["Rounds of iteration"]
                    self.count_try_by = content["Counts of try_by"]
                    self.time_global = content["Global Time"]
                    self.subUCRate = content['SubUCRate']
                    if content.get("Propagation") !=None:
                        prop = content["Propagation"]
                        self.count_prop = prop["Amount"]
                        self.time_prop = prop["Total Time"]
                        self.count_prop_suc = prop["Succeed"]
                        if prop.get("UC distribution")!=None:
                            prop_uc_dist = prop["UC distribution"]
                            self.prop_uc_0_3 = prop_uc_dist["[0,3)"]
                            self.prop_uc_3_5 = prop_uc_dist["[3,5)"]
                            self.prop_uc_5_10 = prop_uc_dist["[5,10)"]
                            self.prop_uc_10_50 = prop_uc_dist["[10,50)"]
                            self.prop_uc_50_100 = prop_uc_dist["[50,100)"]
                            self.prop_uc_100_inf = prop_uc_dist["[100,inf)"]
                        
                    uc_dist = content["UC distribution"]
                    self.uc_0_3 = uc_dist["[0,3)"]
                    self.uc_3_5 = uc_dist["[3,5)"]
                    self.uc_5_10 = uc_dist["[5,10)"]
                    self.uc_10_50 = uc_dist["[10,50)"]
                    self.uc_50_100 = uc_dist["[50,100)"]
                    self.uc_100_inf = uc_dist["[100,inf)"]
        except Exception as e:
            print(f"error dealing with {path}:{e}")
            exit(0)


def process_log(dir):
    """
    deal with logs in this dir. No recursion.
    """
    print(f"processing {dir}")
    result = dict()
    files = os.listdir(dir)
    if len(files) != 337:
        print(f"Warning: missing files {dir}")
    for file in files:
        if os.path.isdir(file):
            # do nothing with subdir
            continue
        else:
            if file.endswith(".log"):
                name = getbasename(file)
                result[name] = info(os.path.join(dir, file))
    result = sorted(
        result.items(),
        key=lambda x: (
            getattr(x[1], "status", None) != "missed",
            getattr(x[1], "status", None) != "timeout",
            getattr(x[1], "time_original", None),
        ),
    )
    return result


def serialize(dataDict, sheet_name, wb):
    # wb = Workbook()
    ws = wb.create_sheet(title=sheet_name)
    header = [
        "name",
        "status",
        "time(global)",
        "time(original)",
        "time(conv)",
        "time(imply)",
        "impDecision",
        "count(original)",
        "count(original_succ)",
        "count(original_fail)",
        "count(tried before)",
        "avg(original uc)",
        "count(conv)",
        "count(shorter uc)",
        "avg(conv uc)",
        "count(iter)",
        "count(try by)",
        "subUCRate",
        "count(Prop)",
        "count(PropSucc)",
        "time(Prop)",
        "uc_0_3",
        "uc_3_5",
        "uc_5_10",
        "uc_10_50",
        "uc_50_100",
        "uc_100_inf",
        "prop_uc_0_3",
        "prop_uc_3_5",
        "prop_uc_5_10",
        "prop_uc_10_50",
        "prop_uc_50_100",
        "prop_uc_100_inf",
    ]
    ws.append(header)
    for index,iter in enumerate(dataDict):
        row = []
        for attr in [
            "name",
            "status",
            "time_global",
            "time_original",
            "time_conv",
            "time_imply",
            "impDecision",
            "count_original",
            "count_original_success",
            "count_original_failed",
            "count_tried_before",
            "count_original_uc_len",
            "count_conv",
            "count_conv_shorter",
            "count_conv_uc_len",
            "count_rounds",
            "count_try_by",
            "subUCRate",
            "count_prop",
            "count_prop_suc",
            "time_prop",
            "uc_0_3",
            "uc_3_5",
            "uc_5_10",
            "uc_10_50",
            "uc_50_100",
            "uc_100_inf",
            "prop_uc_0_3",
            "prop_uc_3_5",
            "prop_uc_5_10",
            "prop_uc_10_50",
            "prop_uc_50_100",
            "prop_uc_100_inf",
        ]:
            item = iter[1]
            if hasattr(item, attr):
                value = getattr(item, attr)
                row.append(value)
            else:
                row.append("missed")
        ws.append(row)
        # 那些新找出的，从来没解出来过的
        if hasattr(item,"color") and item.color=="red":
            cell = ws.cell(row=index+2, column=1,value=iter[1].name)
            cell.font = Font(color="FF0000")  # 设置字体颜色为红色
    # wb.save('log_data.xlsx')


def process_and_print(dir):
    global res
    wb = Workbook()
    recMode = True
    for subdir in os.listdir(dir):
        path = os.path.join(dir, subdir)
        if os.path.isdir(path):
            recMode = True
            print("folded dir, recursion")
            break
        else:
            recMode = False
            print("single dir, no recursion")
            break

    for subdir in os.listdir(dir):
        path = os.path.join(dir, subdir)
        if os.path.isdir(path):
            res = process_log(path)
            serialize(res, os.path.basename(path), wb)
        else:
            if recMode is True:
                continue
            else:
                res = process_log(dir)
                serialize(res, os.path.basename(dir), wb)
                break
    base = os.path.basename(dir).replace("tmp_", "")
    wb.save(f"log_{base}.xlsx")


if __name__ == "__main__":
    path = "./tmp_291e021/291e021/continue"
    process_and_print(path)
    print(f"vbs : {len(vbs)}")
    print(f"breakthrough : {len(vbs.intersection(hard))} : {vbs.intersection(hard)}")
    print(f"breakthrough : {breakthrough}")
    # print(f"vbs : {vbs}")
